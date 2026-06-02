"""
SGM Caloi — Portal de Gestão de Manutenção
Caloi Norte SA · Distrito Industrial de Manaus
Versão 2.0 · 2025

Portal executivo com 11 módulos de indicadores.
Integração direta com o banco do TPM via volume Railway compartilhado.
"""

import os, json, hashlib, base64, shutil, io
import urllib.request, urllib.parse
from pathlib import Path
from datetime import datetime, date, timedelta
from functools import wraps
from flask import (Flask, render_template, request, redirect,
                   url_for, session, jsonify, send_file, abort)

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ─────────────────────────────────────────────────────────────
# CAMINHOS
# ─────────────────────────────────────────────────────────────
APP_DIR  = Path(__file__).parent
DATA_DIR = Path(os.environ.get("DATA_DIR", str(APP_DIR)))

# Dados do SGM (próprios)
SGM_DIR       = DATA_DIR / "sgm"
USUARIOS_FILE = SGM_DIR / "usuarios.json"
UPLOADS_DIR   = DATA_DIR / "uploads"   # Excel de cada módulo

# Dados do TPM (compartilhados — mesmo volume Railway)
def _detect_tpm_dir():
    # 1) Produção (Railway): variável de ambiente apontando para o volume
    env = os.environ.get("TPM_DATA_DIR")
    if env:
        return Path(env)
    # 2) Dev local: pasta irmã "Manutencao_TPM" ao lado do SGM
    sib = APP_DIR.parent / "Manutencao_TPM"
    if (sib / "Manutencao_TPM.xlsx").exists():
        return sib
    # 3) Fallback: própria pasta de dados
    return DATA_DIR

TPM_DIR       = _detect_tpm_dir()
TPM_OS_FILE   = TPM_DIR / "Manutencao_TPM.xlsx"   # arquivo principal do TPM
TPM_PLANO     = TPM_DIR / "Plano_de_Acao.xlsx"
TPM_AGENDA    = TPM_DIR / "Agenda_PM.xlsx"         # se existir
TPM_COLAB     = TPM_DIR / "Colaboradores.xlsx"     # cadastro do efetivo (editável no TPM)

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "sgm-portal-caloi-2025")

# URL do programa TPM (formulário de OS). Configure em produção via env TPM_URL.
# (link usado pelo ícone "OS de Manutenção" e botões "Abrir TPM")
TPM_URL = os.environ.get("TPM_URL", "https://web-production-97918.up.railway.app")
# API do TPM para ler dados online (quando o Excel local não está disponível, ex.: Railway)
TPM_API_URL = os.environ.get("TPM_API_URL", TPM_URL)
SGM_API_KEY = os.environ.get("SGM_API_KEY", "caloi-sgm-2026")  # deve ser igual no TPM

# ─────────────────────────────────────────────────────────────
# AUTENTICAÇÃO
# ─────────────────────────────────────────────────────────────
def hash_senha(s):
    return hashlib.sha256(s.encode()).hexdigest()

# Senha do admin vem de variável de ambiente (defina ADMIN_PASSWORD no Railway).
# Em dev/local, cai no padrão "admin123".
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "admin123")

USUARIOS_PADRAO = [
    {"id":"1","username":"admin","password_hash":hash_senha(ADMIN_PASSWORD),
     "nome":"Administrador","perfil":"admin","ativo":True},
    {"id":"2","username":"gerente","password_hash":hash_senha(os.environ.get("GERENTE_PASSWORD","ger123")),
     "nome":"Gerente","perfil":"gerente","ativo":True},
    {"id":"3","username":"viewer","password_hash":hash_senha(os.environ.get("VIEWER_PASSWORD","view123")),
     "nome":"Visualizador","perfil":"viewer","ativo":True},
]

def carregar_json(path, default):
    try:
        if Path(path).exists():
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return default

def salvar_json(path, dados):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)

def usuario_logado():
    uid = session.get("user_id")
    if not uid:
        return None
    usuarios = carregar_json(USUARIOS_FILE, []) or USUARIOS_PADRAO
    for u in usuarios:
        if u["id"] == uid and u.get("ativo", True):
            return u
    return None

def login_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if not usuario_logado():
            return redirect(url_for("login", next=request.path))
        return f(*a, **kw)
    return dec

@app.context_processor
def inject_usuario():
    return {"usuario": usuario_logado()}

# ─────────────────────────────────────────────────────────────
# INICIALIZAÇÃO
# ─────────────────────────────────────────────────────────────
def init():
    SGM_DIR.mkdir(parents=True, exist_ok=True)
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    if not USUARIOS_FILE.exists():
        salvar_json(USUARIOS_FILE, USUARIOS_PADRAO)

try:
    init()
except Exception as e:
    print(f"[WARN] init: {e}")

# ─────────────────────────────────────────────────────────────
# LEITURA DO EXCEL DO TPM
# ─────────────────────────────────────────────────────────────
def ler_excel_tpm(path):
    """Lê o Excel principal do TPM e retorna lista de dicts."""
    if not HAS_OPENPYXL or not Path(path).exists():
        return []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        headers = [str(c.value or "").strip() for c in next(ws.iter_rows(max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2):
            vals = [c.value for c in row]
            if any(v is not None for v in vals):
                rows.append(dict(zip(headers, vals)))
        wb.close()
        return rows
    except Exception as e:
        print(f"[WARN] ler_excel_tpm: {e}")
        return []

def ler_excel_upload(modulo):
    """Lê o Excel de upload de um módulo específico."""
    path = UPLOADS_DIR / f"{modulo}.xlsx"
    if not path.exists():
        # Tentar .xls
        path = UPLOADS_DIR / f"{modulo}.xls"
    return ler_excel_tpm(path) if path.exists() else []

def serial(v):
    """Serializa valores para JSON."""
    if v is None:
        return None
    if hasattr(v, "isoformat"):
        return str(v)
    if hasattr(v, "__float__"):
        try:
            return float(v)
        except Exception:
            return str(v)
    return str(v) if not isinstance(v, (str, int, float, bool)) else v

def serializar_rows(rows):
    return [{k: serial(v) for k, v in r.items()} for r in rows]

def fetch_tpm_api(dataset):
    """Busca os dados do TPM via HTTP (usado quando não há Excel local)."""
    if not TPM_API_URL:
        return None
    try:
        url = f"{TPM_API_URL.rstrip('/')}/api/dados/{dataset}?key={urllib.parse.quote(SGM_API_KEY)}"
        with urllib.request.urlopen(url, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        return data if isinstance(data, list) else None
    except Exception as e:
        print(f"[WARN] fetch_tpm_api({dataset}): {e}")
        return None

def carregar_dataset(dataset, local_path):
    """Lê do Excel local (dev). Se vazio/ausente, busca via API do TPM (produção)."""
    rows = ler_excel_tpm(local_path)
    if rows:
        return rows
    api = fetch_tpm_api(dataset)
    return api or []

# ─────────────────────────────────────────────────────────────
# KPIs DO TPM
# ─────────────────────────────────────────────────────────────
def _num(v):
    try:
        return float(str(v).replace(",", "."))
    except Exception:
        return 0.0

def calcular_kpis_tpm(registros):
    """KPIs a partir do esquema real do TPM (registro de paradas).
    Colunas: Status, Total Horas Paradas, Equipamento, Setor Produtivo, etc."""
    total = len(registros)
    if not total:
        return {"total":0,"concluidas":0,"abertas":0,"corretivas":0,
                "preventivas":0,"mttr":0,"mtbf":0,"disponibilidade":0,
                "horas_paradas":0,"equipamentos":0}
    conc = sum(1 for r in registros if "conclu" in str(r.get("Status","")).lower())
    horas = [_num(r.get("Total Horas Paradas", 0)) for r in registros]
    com_horas = [h for h in horas if h > 0]
    tt = sum(com_horas)                       # total de horas paradas
    mttr = round(tt/len(com_horas), 2) if com_horas else 0       # horas/OS
    mtbf = round(720/total, 1) if total else 0                   # h (janela ~mensal)
    disp = round(max(0, min(100, (1 - tt/720) * 100)), 1)        # disponibilidade %
    equipamentos = len({str(r.get("Equipamento","")).strip()
                        for r in registros if r.get("Equipamento")})
    return {"total":total,"concluidas":conc,"abertas":total-conc,
            "corretivas":0,"preventivas":0,"mttr":mttr,"mtbf":mtbf,
            "disponibilidade":disp,"horas_paradas":round(tt,1),
            "equipamentos":equipamentos}

def resumo_modulos():
    """Gera mini-resumo para cada card do portal."""
    os_data = carregar_dataset("os", TPM_OS_FILE)
    kpis    = calcular_kpis_tpm(os_data)
    plano   = carregar_dataset("plano", TPM_PLANO)
    abertas_plano = sum(1 for r in plano if str(r.get("Status","")).lower() not in ("concluída","cancelada"))

    def n_upload(mod):
        p = UPLOADS_DIR / f"{mod}.xlsx"
        if p.exists():
            rows = ler_excel_tpm(p)
            return len(rows)
        return None

    return {
        "ordens":        {"total": kpis["total"], "abertas": kpis["abertas"],
                          "mttr": kpis["mttr"], "disponibilidade": kpis["disponibilidade"]},
        "plano_acao":    {"total": len(plano), "abertas": abertas_plano},
        "custos":        {"registros": n_upload("custos")},
        "pecas":         {"registros": n_upload("pecas")},
        "projetos":      {"registros": n_upload("projetos")},
        "melhoria":      {"registros": n_upload("melhoria")},
        "seguranca":     {"registros": n_upload("seguranca")},
        "headcount":     {"registros": n_upload("headcount")},
        "documentos":    {"registros": n_upload("documentos")},
        "kpi":           {"mttr": kpis["mttr"], "mtbf": kpis["mtbf"],
                          "disponibilidade": kpis["disponibilidade"]},
        "planejamento":  {"registros": n_upload("planejamento")},
    }

# ─────────────────────────────────────────────────────────────
# ROTAS — LOGIN
# ─────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return redirect(url_for("portal"))

@app.route("/login", methods=["GET","POST"])
def login():
    erro = None
    if request.method == "POST":
        username = request.form.get("username","").strip()
        senha    = request.form.get("senha","")
        sh       = hash_senha(senha)
        usuarios = carregar_json(USUARIOS_FILE, []) or USUARIOS_PADRAO
        for u in usuarios:
            if u["username"]==username and u["password_hash"]==sh and u.get("ativo",True):
                session["user_id"] = u["id"]
                session["nome"]    = u["nome"]
                session["perfil"]  = u["perfil"]
                nxt = request.form.get("next","") or request.args.get("next","")
                if nxt and nxt.startswith("/") and not nxt.startswith("//"):
                    return redirect(nxt)
                return redirect(url_for("portal"))
        erro = "Usuário ou senha incorretos."
    return render_template("login.html", erro=erro)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ─────────────────────────────────────────────────────────────
# ROTA — PORTAL PRINCIPAL
# ─────────────────────────────────────────────────────────────
@app.route("/portal")
@login_required
def portal():
    # A capa (roda + peças) é a própria rosto.html, servida diretamente.
    return send_file(str(APP_DIR / "rosto.html"))

@app.route("/fotos/<path:fn>")
def fotos(fn):
    from flask import send_from_directory
    return send_from_directory(str(APP_DIR / "fotos"), fn)

# ─────────────────────────────────────────────────────────────
# ROTAS — MÓDULOS
# ─────────────────────────────────────────────────────────────

@app.route("/modulo/ordens")
@login_required
def mod_ordens():
    rows = carregar_dataset("os", TPM_OS_FILE)
    plano = carregar_dataset("plano", TPM_PLANO)
    kpis = calcular_kpis_tpm(rows)
    anos = sorted({str(r.get("Data",""))[:4] for r in rows if r.get("Data") and str(r.get("Data",""))[:4].isdigit()}, reverse=True)
    return render_template("modulos/ordens.html",
                           kpis=kpis, records_json=serializar_rows(rows),
                           plano_json=serializar_rows(plano),
                           anos=anos, tpm_url=TPM_URL)

@app.route("/modulo/plano-acao")
@login_required
def mod_plano():
    rows  = carregar_dataset("plano", TPM_PLANO)
    return render_template("modulos/plano_acao.html",
                           records_json=serializar_rows(rows),
                           total=len(rows), tpm_url=TPM_URL)

@app.route("/modulo/historico")
@login_required
def mod_historico():
    # Histórico por TAG/equipamento, a partir das OS do TPM.
    rows = carregar_dataset("os", TPM_OS_FILE)
    return render_template("modulos/historico.html",
                           records_json=serializar_rows(rows), tpm_url=TPM_URL)

@app.route("/modulo/custos")
@login_required
def mod_custos():
    rows = ler_excel_upload("custos")
    return render_template("modulos/custos.html",
                           dados=serializar_rows(rows), tem_dados=bool(rows))

@app.route("/modulo/pecas")
@login_required
def mod_pecas():
    # Peças necessárias vêm do Plano de Ação do TPM (campo "Peças Necessárias").
    rows = carregar_dataset("plano", TPM_PLANO)
    return render_template("modulos/pecas.html",
                           records_json=serializar_rows(rows), tpm_url=TPM_URL)

@app.route("/modulo/projetos")
@login_required
def mod_projetos():
    rows = ler_excel_upload("projetos")
    return render_template("modulos/projetos.html",
                           dados=serializar_rows(rows), tem_dados=bool(rows))

@app.route("/modulo/melhoria")
@login_required
def mod_melhoria():
    rows = ler_excel_upload("melhoria")
    return render_template("modulos/melhoria.html",
                           dados=serializar_rows(rows), tem_dados=bool(rows))

@app.route("/modulo/seguranca")
@login_required
def mod_seguranca():
    rows = ler_excel_upload("seguranca")
    return render_template("modulos/seguranca.html",
                           dados=serializar_rows(rows), tem_dados=bool(rows))

@app.route("/modulo/headcount")
@login_required
def mod_headcount():
    # Horas/atividades vêm das OS do TPM; função/férias/capacitação do cadastro.
    rows  = carregar_dataset("os", TPM_OS_FILE)
    colab = carregar_dataset("colaboradores", TPM_COLAB)   # Colaboradores.xlsx
    return render_template("modulos/headcount.html",
                           records_json=serializar_rows(rows),
                           colaboradores=serializar_rows(colab), tpm_url=TPM_URL)

@app.route("/modulo/documentos")
@login_required
def mod_documentos():
    rows = ler_excel_upload("documentos")
    return render_template("modulos/documentos.html",
                           dados=serializar_rows(rows), tem_dados=bool(rows))

@app.route("/modulo/kpi")
@login_required
def mod_kpi():
    rows = carregar_dataset("os", TPM_OS_FILE)
    return render_template("modulos/kpi.html",
                           records_json=serializar_rows(rows), tpm_url=TPM_URL)

@app.route("/modulo/planejamento")
@login_required
def mod_planejamento():
    rows = ler_excel_upload("planejamento")
    # Tentar também Agenda do TPM
    agenda_tpm = ler_excel_tpm(TPM_AGENDA) if TPM_AGENDA.exists() else []
    return render_template("modulos/planejamento.html",
                           dados=serializar_rows(rows),
                           agenda_tpm=serializar_rows(agenda_tpm),
                           tem_dados=bool(rows or agenda_tpm))

# ─────────────────────────────────────────────────────────────
# API — UPLOAD DE EXCEL POR MÓDULO
# ─────────────────────────────────────────────────────────────
@app.route("/api/upload/<modulo>", methods=["POST"])
@login_required
def upload_modulo(modulo):
    MODULOS_VALIDOS = {"custos","pecas","projetos","melhoria","seguranca",
                       "headcount","documentos","planejamento"}
    if modulo not in MODULOS_VALIDOS:
        return jsonify({"success":False,"message":"Módulo inválido"}), 400
    f = request.files.get("arquivo")
    if not f:
        return jsonify({"success":False,"message":"Nenhum arquivo enviado"}), 400
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    dest = UPLOADS_DIR / f"{modulo}.xlsx"
    f.save(str(dest))
    # Contar linhas
    rows = ler_excel_tpm(dest)
    return jsonify({"success":True,
                    "message":f"✅ {len(rows)} registros importados!",
                    "registros": len(rows)})

# ─────────────────────────────────────────────────────────────
# PWA MANIFEST
# ─────────────────────────────────────────────────────────────
@app.route("/manifest.json")
def manifest():
    from flask import send_from_directory
    return send_from_directory(str(APP_DIR / "static"), "manifest.json",
                               mimetype="application/manifest+json")

# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    # SGM usa a porta 5001 (o TPM usa a 5000) — assim os dois rodam juntos.
    port = int(os.environ.get("PORT", 5001))
    app.run(debug=False, host="0.0.0.0", port=port)
